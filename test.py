import pandas as pd 
from openpyxl import load_workbook
import seaborn as sns
import matplotlib.pyplot as plt
import numpy as np



def parse_telecom_data_to_ml_format(file_path, sheet_name="By BDT region") :

    variables = [
        "Fixed-telephone subscriptions/Millions",
        "Fixed-broadband subscriptions/Millions",
        "Mobile-cellular telephone subscriptions/Millions",
        "Active mobile-broadband subscriptions/Millions",
        "Population covered by a mobile-cellular network/Millions",
        "Population covered by at least a 3G mobile network/Millions",
        "Population covered by at least an LTE/WiMAX mobile network/Millions",
        "Population covered by at least a 5G mobile network/Millions",
        "Fixed broadband traffic(Exabytes)",
        "Mobile broadband traffic(Exabytes)",
        "International bandwidth usage(Tbits/s**)",
        "Individuals using the Internet(Millions)",
        "Individuals owning a mobile phone***(Millions)"
    ]

    regions = ["Africa", "Americas", "Europe", "Asia-Pacific", "Arab States", "CIS"]

    years = list(range(2005, 2025))

    try:
        wb = load_workbook(file_path, data_only=True)
        sheet = wb[sheet_name] 

        all_data = []

        current_var = None

        for row in sheet.iter_rows(min_col=1, max_col=50):
            cell_value = row[0].value
            
            if cell_value is None:
                continue
                
            if any(var in str(cell_value) for var in variables):
                current_var = str(cell_value)
                continue
                
            if cell_value in regions and current_var:
                year_values = []
                for i in range(1, len(years) + 1): 
                    if i < len(row) and row[i].value is not None:
                        year_values.append(row[i].value)
                    else:
                        year_values.append(np.nan)
                
                # Create entries for each year
                for year_idx, year in enumerate(years):
                    if year_idx < len(year_values):
                        all_data.append({
                            'Region': cell_value,
                            'Year': year,
                            'Variable': current_var,
                            'Value': year_values[year_idx]
                        })
        
        # Convert to DataFrame
        df_long = pd.DataFrame(all_data)
        
        # Pivot to get variables as columns
        df_ml = df_long.pivot_table(
            index=['Region', 'Year'],
            columns='Variable', 
            values='Value',
            aggfunc='first'
        ).reset_index()
        
        # Clean column names
        df_ml.columns.name = None
        
        # Rename columns to shorter names
        column_mapping = {
            "Fixed-telephone subscriptions/Millions": "Fixed_telephone",
            "Fixed-broadband subscriptions/Millions": "Fixed_broadband",
            "Mobile-cellular telephone subscriptions/Millions": "Mobile_cellular_sub",
            "Active mobile-broadband subscriptions/Millions": "Active_mobile_broadband_sub",
            "Population covered by a mobile-cellular network/Millions": "Coverage_mobile_cellular",
            "Population covered by at least a 3G mobile network/Millions": "Coverage_3G",
            "Population covered by at least an LTE/WiMAX mobile network/Millions": "Coverage_LTE_WiMAX",
            "Population covered by at least a 5G mobile network/Millions": "Coverage_5G",
            "Fixed broadband traffic(Exabytes)": "Fixed_broadband_traffic_EB",
            "Mobile broadband traffic(Exabytes)": "Mobile_broadband_traffic_EB",
            "International bandwidth usage(Tbits/s**)": "Intl_bandwidth_Tbps",
            "Individuals using the Internet(Millions)": "Internet_users",
            "Individuals owning a mobile phone***(Millions)": "Mobile_phone_owners"
        }
        
        df_ml = df_ml.rename(columns=column_mapping)
        
        # Sort by Region and Year
        df_ml = df_ml.sort_values(['Region', 'Year']).reset_index(drop=True)
        
        return df_ml
        
    except FileNotFoundError:
        print(f"Error: File '{file_path}' not found.")
        return None
    except Exception as e:
        print(f"Error loading data: {e}")
        return None


def load_data(file_path, vars, regions) :

    """
    Load telecommunications data from Excel file
    Returns two DataFrames: absolute values and relative values
    """

    data_absolute = []
    data_relative = []

    try:

        wb = load_workbook(file_path, data_only=True)
        sheet = wb["By BDT region"]

        current_var = None

        for row in sheet.iter_rows(min_col=1, max_col=50):  

            cell_value = row[0].value
            
            if cell_value is None:
                continue

            if cell_value in vars:
                current_var = cell_value
                continue

            if cell_value in regions and current_var:

                #Absolute Val
                year_values_abs = []
                for i in range(1, 21):  
                    if i < len(row):
                        year_values_abs.append(row[i].value)
                    else:
                        year_values_abs.append(None)

                entry_abs = {"Var": current_var, "Region": cell_value}
                for i, year in enumerate(range(2005, 2025)):
                    if i < len(year_values_abs):
                        entry_abs[str(year)] = year_values_abs[i]
                    else:
                        entry_abs[str(year)] = None
                data_absolute.append(entry_abs)

                #Relative val
                year_values_rel = []
                for i in range(22, 42): 
                    if i < len(row):
                        year_values_rel.append(row[i].value)
                    else:
                        year_values_rel.append(None)
                
                entry_rel = {"Var": current_var, "Region": cell_value}
                for i, year in enumerate(range(2005, 2025)):                
                    if i < len(year_values_rel):
                        entry_rel[str(year)] = year_values_rel[i]
                    else:
                        entry_rel[str(year)] = None
                data_relative.append(entry_rel)

    except FileNotFoundError:
        print(f"Error: File '{file_path}' not found.")
    
    except Exception as e:
        print(f"Error loading data: {e}")

    return data_absolute, data_relative