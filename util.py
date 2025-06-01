"""
Notes:																																													
* 1 Exabyte = 1012 Megabytes																																													
** 1 terabit = 1’000’000 megabits																																													
*** Individuals aged 10 or older																																													
All data are ITU estimates																																													
Values are rounded to 1 decimal																																													
N/A: Not available																																													
Regions are based on the ITU regions, see: http://www.itu.int/en/ITU-D/Statistics/Pages/definitions/regions.aspx 																																													
																																													
Source:																																													
ITU World Telecommunication/ICT Indicators database																																													
Version November 2024, for Facts and Figures 2024																																													

"""

import pandas as pd 
from openpyxl import load_workbook
import seaborn as sns
import numpy as np 
import matplotlib.pyplot as plt
from scipy import stats
import warnings
warnings.filterwarnings('ignore')


def load_data(file_path, sheet_name="By BDT region") :

    variables = [
        "Fixed-telephone subscriptions/Millions",
        "Fixed-broadband subscriptions/Millions",
        "Mobile-cellular telephone subscriptions/Millions",
        "Active mobile-broadband subscriptions/Millions",
        "Population covered by a mobile-cellular network/Millions",
        "Population covered by at least a 3G mobile network/Millions",
        "Population covered by at least an LTE/WiMAX mobile network/Millions",
        "Population covered by at least a 5G mobile network/Millions",
        "Fixed broadband traffic(Exabytes)",
        "Mobile broadband traffic(Exabytes)",
        "International bandwidth usage(Tbits/s**)",
        "Individuals using the Internet(Millions)",
        "Individuals owning a mobile phone***(Millions)"
    ]

    regions = ["Africa", "Americas", "Europe", "Asia-Pacific", "Arab States", "CIS"]

    years = list(range(2005, 2025))

    try:
        wb = load_workbook(file_path, data_only=True)
        sheet = wb[sheet_name] 

        abs_data = []
        rel_data = []

        current_var = None

        for row in sheet.iter_rows(min_col=1, max_col=50):
            cell_value = row[0].value

            if cell_value is None:
                continue
                
            if any(var in str(cell_value) for var in variables):
                current_var = str(cell_value)
                continue
                
            if cell_value in regions and current_var:

                """abs"""
                year_values = []
                for i in range(1, len(years) + 1): 
                    if i < len(row) and row[i].value is not None:
                        year_values.append(row[i].value)
                    else:
                        year_values.append(np.nan)
                
                # Create entries for each year
                for year_idx, year in enumerate(years):
                    if year_idx < len(year_values):
                        abs_data.append({
                            'Region': cell_value,
                            'Year': year,
                            'Variable': current_var,
                            'Value': year_values[year_idx]
                        })

                """Relative"""
                year_values = []
                for i in range(len(years) +2, len(years) + 24): 
                    if i < len(row) and row[i].value is not None:
                        year_values.append(row[i].value)
                    else:
                        year_values.append(np.nan)
                
                # Create entries for each year
                for year_idx, year in enumerate(years):
                    if year_idx < len(year_values):
                        rel_data.append({
                            'Region': cell_value,
                            'Year': year,
                            'Variable': current_var,
                            'Value': year_values[year_idx]
                        })
        
        # Convert to DataFrame
        df_long_abs = pd.DataFrame(abs_data)
        df_long_rel = pd.DataFrame(rel_data)

        
        # Pivot to get variables as columns
        df_ml_abs = df_long_abs.pivot_table(
            index=['Region', 'Year'],
            columns='Variable', 
            values='Value',
            aggfunc='first'
        ).reset_index()

        # Pivot as usual
        df_ml_rel = df_long_rel.pivot_table(
            index=['Region', 'Year'],
            columns='Variable', 
            values='Value',
            aggfunc='first'
        ).reset_index()

        
        # Clean column names
        df_ml_abs.columns.name = None
        df_ml_rel.columns.name = None
        
        # Rename columns to shorter names
        column_mapping = {
            "Fixed-telephone subscriptions/Millions": "Fixed_telephone",
            "Fixed-broadband subscriptions/Millions": "Fixed_broadband",
            "Mobile-cellular telephone subscriptions/Millions": "Mobile_cellular_sub",
            "Active mobile-broadband subscriptions/Millions": "Active_mobile_broadband_sub",
            "Population covered by a mobile-cellular network/Millions": "Coverage_mobile_cellular",
            "Population covered by at least a 3G mobile network/Millions": "Coverage_3G",
            "Population covered by at least an LTE/WiMAX mobile network/Millions": "Coverage_LTE_WiMAX",
            "Population covered by at least a 5G mobile network/Millions": "Coverage_5G",
            "Fixed broadband traffic(Exabytes)": "Fixed_broadband_traffic_EB",
            "Mobile broadband traffic(Exabytes)": "Mobile_broadband_traffic_EB",
            "International bandwidth usage(Tbits/s**)": "Intl_bandwidth_Tbps",
            "Individuals using the Internet(Millions)": "Internet_users",
            "Individuals owning a mobile phone***(Millions)": "Mobile_phone_owners"
        }
        
        df_ml_abs = df_ml_abs.rename(columns=column_mapping)
        df_ml_rel = df_ml_rel.rename(columns=column_mapping)
        
        # Sort by Region and Year
        df_ml_abs = df_ml_abs.sort_values(['Region', 'Year']).reset_index(drop=True)
        df_ml_rel = df_ml_rel.sort_values(['Region', 'Year']).reset_index(drop=True)
        
        return df_ml_abs, df_ml_rel
        
    except FileNotFoundError:
        print(f"Error: File '{file_path}' not found.")
        return None
    except Exception as e:
        print(f"Error loading data: {e}")
        return None
    

def prepare_initial_data(df, variable_columns=None):
    # Ensure the data is in the correct format
    if variable_columns is None:
        variable_columns = [
            "Fixed_telephone", "Fixed_broadband", "Mobile_cellular_sub", 
            "Active_mobile_broadband_sub", "Coverage_mobile_cellular", 
            "Coverage_3G", "Coverage_LTE_WiMAX", "Coverage_5G", 
            "Fixed_broadband_traffic_EB", "Mobile_broadband_traffic_EB", 
            "Intl_bandwidth_Tbps", "Internet_users", "Mobile_phone_owners"
        ]
    
    df_clean = df.copy()
    for col in variable_columns:
        if col in df_clean.columns:
            df_clean[col] = pd.to_numeric(df_clean[col], errors='coerce')
    
    return df_clean

def explore_dataframe(df, name="DataFrame", save_plots=False):
    
    #EDA

    print(f"\n{'='*60}")
    print(f"EDA REPORT FOR {name.upper()}")
    print(f"{'='*60}")

    # 1. Basic Info
    print(f"\n1. DATASET OVERVIEW")
    print(f"   Shape: {df.shape[0]} rows x {df.shape[1]} columns")
    print(f"   Date range: {df['Year'].min()} - {df['Year'].max()}")
    print(f"   Regions: {', '.join(df['Region'].unique())}")

    # 2. Descriptive Statistics
    print(f"\n2. DESCRIPTIVE STATISTICS")
    numeric_cols = df.select_dtypes(include='number').columns
    desc_stats = df[numeric_cols].describe()
    print(desc_stats)

    # 3. Missing Values Analysis
    print(f"\n3. MISSING VALUES ANALYSIS")
    missing = df.isnull().sum()
    missing_pct = (missing / len(df)) * 100
    missing_df = pd.DataFrame({
        'Missing_Count': missing,
        'Missing_Percentage': missing_pct
    }).sort_values('Missing_Count', ascending=False)
    
    missing_vars = missing_df[missing_df['Missing_Count'] > 0]
    if not missing_vars.empty:
        print(missing_vars)
        
        # Missing values heatmap
        plt.figure(figsize=(14, 8))
        sns.heatmap(df.isnull(), cbar=True, yticklabels=False, cmap='viridis')
        plt.title(f"Missing Values Heatmap - {name}")
        plt.xlabel("Variables")
        plt.tight_layout()
        if save_plots: plt.savefig(f'missing_values_{name.lower()}.png', dpi=300, bbox_inches='tight')
        plt.show()
    else:
        print("No missing values found!")

    # 4. Data Quality Assessment
    print(f"\n4. DATA QUALITY ASSESSMENT")
    
    # Check for negative values in variables that shouldn't have them
    negative_vars = []
    for col in numeric_cols:
        if col != 'Year' and (df[col] < 0).any():
            negative_count = (df[col] < 0).sum()
            negative_vars.append(f"{col}: {negative_count} negative values")
    
    if negative_vars:
        print("Negative values found:")
        for var in negative_vars:
            print(f"   {var}")
    else:
        print("No unexpected negative values found")

    # 5. Distribution Analysis
    print(f"\n5. DISTRIBUTION ANALYSIS")
    
    # Plot distributions for key variables
    key_vars = [col for col in ['Mobile_cellular_sub', 'Internet_users', 'Fixed_broadband', 'Coverage_mobile_cellular'] 
                if col in df.columns]
    
    if key_vars:
        fig, axes = plt.subplots(2, 2, figsize=(15, 10))
        axes = axes.flatten()
        
        for i, var in enumerate(key_vars[:4]):
            if i < len(axes):
                df[var].dropna().hist(bins=30, ax=axes[i], alpha=0.7, edgecolor='black')
                axes[i].set_title(f'Distribution of {var}')
                axes[i].set_xlabel(var)
                axes[i].set_ylabel('Frequency')
        
        plt.tight_layout()
        if save_plots: plt.savefig(f'distributions_{name.lower()}.png', dpi=300, bbox_inches='tight')
        plt.show()

    # 6. Correlation Analysis
    if len(numeric_cols) > 2:
        print(f"\n6. CORRELATION ANALYSIS")
        
        corr_matrix = df[numeric_cols].corr()
        
        # Find highly correlated pairs
        high_corr_pairs = []
        for i in range(len(corr_matrix.columns)):
            for j in range(i+1, len(corr_matrix.columns)):
                corr_val = abs(corr_matrix.iloc[i, j])
                if corr_val > 0.8 and not np.isnan(corr_val):
                    high_corr_pairs.append((
                        corr_matrix.columns[i], 
                        corr_matrix.columns[j], 
                        corr_matrix.iloc[i, j]
                    ))
        
        print(f"High correlations (|r| > 0.8):")
        for var1, var2, corr_val in sorted(high_corr_pairs, key=lambda x: abs(x[2]), reverse=True)[:10]:
            print(f"   {var1} ↔ {var2}: {corr_val:.3f}")
        
        # Correlation heatmap
        plt.figure(figsize=(16, 12))
        mask = np.triu(np.ones_like(corr_matrix, dtype=bool))
        sns.heatmap(corr_matrix, annot=True, fmt=".2f", cmap="RdBu_r", 
                   center=0, mask=mask, square=True, cbar_kws={"shrink": .8})
        plt.title(f"Correlation Matrix - {name}")
        plt.tight_layout()
        if save_plots: plt.savefig(f'correlation_{name.lower()}.png', dpi=300, bbox_inches='tight')
        plt.show()

    # 7. Regional Analysis
    print(f"\n7. REGIONAL ANALYSIS")
    
    # Latest year summary by region
    latest_year = df['Year'].max()
    latest_data = df[df['Year'] == latest_year]
    
    print(f"{latest_year} Regional Summary:")
    key_metrics = ['Mobile_cellular_sub', 'Internet_users', 'Fixed_broadband']
    available_metrics = [col for col in key_metrics if col in df.columns]
    
    if available_metrics:
        for metric in available_metrics:
            print(f"\n   {metric}:")
            metric_data = latest_data[['Region', metric]].dropna().sort_values(metric, ascending=False)
            for _, row in metric_data.head(3).iterrows():
                print(f"      🏆 {row['Region']}: {row[metric]:.1f}")

    # 8. Time Series Analysis
    print(f"\n8. TIME SERIES TRENDS")
    
    # Calculate growth rates
    growth_analysis = {}
    for region in df['Region'].unique():
        region_data = df[df['Region'] == region].sort_values('Year')
        for col in numeric_cols:
            if col != 'Year' and col in region_data.columns:
                first_val = region_data[col].dropna().iloc[0] if not region_data[col].dropna().empty else None
                last_val = region_data[col].dropna().iloc[-1] if not region_data[col].dropna().empty else None
                
                if first_val and last_val and first_val > 0:
                    years_span = region_data['Year'].max() - region_data['Year'].min()
                    cagr = ((last_val / first_val) ** (1/years_span) - 1) * 100
                    growth_analysis[f"{region}_{col}"] = cagr

    # Show top growth rates
    if growth_analysis:
        print(f"\nHighest Growth Rates (CAGR %):")
        sorted_growth = sorted(growth_analysis.items(), key=lambda x: x[1], reverse=True)
        for metric, growth_rate in sorted_growth[:10]:
            if not np.isnan(growth_rate) and not np.isinf(growth_rate):
                region, var = metric.split('_', 1)
                print(f"   📈 {region} - {var}: {growth_rate:.1f}% CAGR")

    # 9. Time Series Visualizations
    print(f"\n9. GENERATING TIME SERIES PLOTS...")
    
    id_vars = ['Region', 'Year']
    value_vars = [col for col in df.columns if col not in id_vars and df[col].dtype in ['float64', 'int64']]
    
    # Create subplot layout for multiple variables
    n_vars = len(value_vars)
    n_cols = 2
    n_rows = (n_vars + n_cols - 1) // n_cols
    
    if n_vars > 0:
        fig, axes = plt.subplots(n_rows, n_cols, figsize=(20, 5*n_rows))
        if n_rows == 1:
            axes = [axes] if n_cols == 1 else axes
        else:
            axes = axes.flatten()
        
        for i, var in enumerate(value_vars):
            if i < len(axes):
                ax = axes[i]
                for region in df['Region'].unique():
                    region_df = df[df['Region'] == region].sort_values('Year')
                    if not region_df[var].dropna().empty:
                        ax.plot(region_df['Year'], region_df[var], 
                               label=region, marker='o', linewidth=2, markersize=4)
                
                ax.set_title(f"{var} Over Time by Region", fontsize=12, fontweight='bold')
                ax.set_xlabel("Year")
                ax.set_ylabel(var.replace('_', ' ').title())
                ax.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
                ax.grid(True, alpha=0.3)
        
        # Hide empty subplots
        for i in range(n_vars, len(axes)):
            axes[i].set_visible(False)
        
        plt.tight_layout()
        if save_plots: plt.savefig(f'timeseries_{name.lower()}.png', dpi=300, bbox_inches='tight')
        plt.show()

    # 10. Summary Insights
    print(f"\n 10. KEY INSIGHTS SUMMARY")
    print(f"    Dataset contains {df.shape[0]} observations across {len(df['Region'].unique())} regions")
    print(f"    Time span: {df['Year'].max() - df['Year'].min() + 1} years ({df['Year'].min()}-{df['Year'].max()})")
    
    if missing_vars.empty:
        print(f"    Complete data coverage")
    else:
        print(f"     {len(missing_vars)} variables have missing data")
    
    print(f"    Found {len([x for x in high_corr_pairs if abs(x[2]) > 0.9])} very high correlations (>0.9)")
    print(f"\n{'='*60}")


